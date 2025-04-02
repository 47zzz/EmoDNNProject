import csv
from openpyxl import Workbook

            # 創建一個新的Excel工作簿
wb = Workbook()
ws = wb.active

# 循环处理每个CSV文件

for i in range(1, 36):  # 从ED1到ED35
    csv_filename = f'/Users/477z/Desktop/EmoDNN/behavior/RAWratingData/ED{i}.csv'  # 构造CSV文件名
    with open(csv_filename, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        
        # 遍歷每一行
        mouse_x_values = []
        mouse_y_values = []

        for row in reader:
            # 如果行名為"mouse.x"，則印出該行的值
            if 'mouse.x' in row:
                mouse_x_values.append(row['mouse.x'])
            if 'mouse.y' in row:
                mouse_y_values.append(row['mouse.y'])
                
        for j, value in enumerate(mouse_x_values, start=1):
            ws.cell(row=j, column=i*2, value=value)

        for j, value in enumerate(mouse_y_values, start=1):
            ws.cell(row=j, column=i*2 + 1, value=value)
# 將值寫入Excel文件

csv_filename = f'/Users/477z/Desktop/EmoDNN/behavior/stimuli_list.csv'  # 构造CSV文件名
with open(csv_filename, newline='') as csvfile:
    reader = csv.DictReader(csvfile)
    
    # 遍歷每一行
    col = []

    for row in reader:
        # 如果行名為"mouse.x"，則印出該行的值
        if '0' in row:
            col.append(row['0'])
            

    for j, value in enumerate(col, start=1):
        ws.cell(row=j, column=1, value=value)

# 儲存Excel文件
wb.save("cleanData.xlsx")
#裡面第一colume為圖片名稱 而後每兩個一組 為同一個受試者的滑鼠x與y 依序由受試者編號1~35

